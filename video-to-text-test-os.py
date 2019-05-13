# USAGE
# python text_detection_video.py --east frozen_east_text_detection.pb



# import the necessary packages
from imutils.video import FPS
from imutils.object_detection import non_max_suppression
import numpy as np
import time
import cv2
import pytesseract
from PIL import Image
from pytesseract import Output
import httplib2
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from threading import Thread
import difflib
from difflib import SequenceMatcher
from collections import deque
import  os
import csv



def initOutputSheet():
	CREDENTIALS_FILE = 'video-to-text-7c0eb4e245da.json'  # имя файла с закрытым ключом

	credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE,
																   ['https://www.googleapis.com/auth/spreadsheets',
																	'https://www.googleapis.com/auth/drive'])
	httpAuth = credentials.authorize(httplib2.Http())

	# создание Книги
	service = build('sheets', 'v4', http=httpAuth)
	spreadsheet = service.spreadsheets().create(body={
		'properties': {'title': 'Output', 'locale': 'ru_RU'},
		'sheets': [{'properties': {'sheetType': 'GRID',
								   'sheetId': 0,
								   'title': 'Tmp',
								   'gridProperties':{'rowCount':10000,'columnCount':100}}},
				   {'properties': {'sheetType': 'GRID',
								   'sheetId': 1,
								   'title': 'Result',
								   'gridProperties': {'rowCount':10000,'columnCount': 100}}}
				   ]
	}).execute()
	driveService = build('drive', 'v3', http=httpAuth)
	shareRes = driveService.permissions().create(
		fileId=spreadsheet['spreadsheetId'],
		body={'type': 'anyone', 'role': 'writer'},  # доступ на чтение кому угодно
		fields='id'
	).execute()
	return service,spreadsheet

def writeToSheet(text,spreadsheet,service,frameNum,time):
	global responseCount
	data = []
	tmp_dict ={}
	tmp_dict['range'] = 'Tmp!' + 'A' + str(frameNum)
	tmp_dict['majorDimension'] = "ROWS"
	tmp_dict['values'] = [[str(time)]]
	data.append(tmp_dict)
	for line in text:
		tmp_dict={}
		tmp_dict['range'] ='Tmp!'+get_column_letter(int(line['Zone'])+1)+str(frameNum)
		tmp_dict['majorDimension']="ROWS"
		tmp_dict['values']=[[str(line['Text'])]]
		data.append(tmp_dict)
	# Заполнение данными
	request = service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet['spreadsheetId'], body={
		"valueInputOption": "USER_ENTERED",
		"data": data
	})
	responseCount += 1
	queue.append({'Number': responseCount, 'Request': request, 'Type': 'set'})
	return

def GoogleRequestQueue():
	global listen
	while listen:
		global queue,googleSheetsService,responseStack,responseCount
		if(len(queue)!=0):
			elem = queue.popleft()
			request=elem['Request']
			requestNumber=elem['Number']
			rtype = elem['Type']
			response =request.execute()
			if rtype == 'get':
				responseStack.append({'Number':requestNumber,'Response':response})
	return



#levenstein
def distance(a, b):
    "Calculates the Levenshtein distance between a and b."
    n, m = len(a), len(b)
    if n > m:
        # Make sure n <= m, to use O(min(n,m)) space
        a, b = b, a
        n, m = m, n

    current_row = range(n+1) # Keep current and previous row, not entire matrix
    for i in range(1, m+1):
        previous_row, current_row = current_row, [i]+[0]*n
        for j in range(1,n+1):
            add, delete, change = previous_row[j]+1, current_row[j-1]+1, previous_row[j-1]
            if a[j-1] != b[i-1]:
                change += 1
            current_row[j] = min(add, delete, change)

    return current_row[n]


def ParseTessData():
	with open('output.tsv') as tsvfile:
		minconf = 40
		content = ''
		reader = csv.DictReader(tsvfile, dialect='excel-tab')
		for row in reader:
			if (int(row['conf']) >= minconf):
				content += ' ' + row['text']
	return content


def SquareInteractPercent(zstartX,zstartY,zendX,zendY,cstartX,cstartY,cendX,cendY):
	#if (zstartX >= cendX) |(cstartX >= zendX) | (cendY >= zstartY) | (zendY >= cstartY):
	if (zstartX >= cendX) | (cstartX >= zendX) | (zstartY>=cendY)|(cstartY>=zendY):
		return 0.0
	else:
		X=[]
		X.append(zstartX)
		X.append(cstartX)
		X.append(zendX)
		X.append(cendX)
		X.sort()
		Y = []
		Y.append(zstartY)
		Y.append(cstartY)
		Y.append(zendY)
		Y.append(cendY)
		Y.sort()
		S_INTERRACT = (X[2]-X[1])*(Y[2]-Y[1])
		S_Z = (zendX-zstartX)*(zendY-zstartY)
		S_C = (cendX - cstartX) * (cendY - cstartY)
		S = min(S_Z,S_C)
		if (S!=0):
			return(S_INTERRACT/S)
		else:
			return  0


def decode_predictions(scores, geometry,args):
	# grab the number of rows and columns from the scores volume, then
	# initialize our set of bounding box rectangles and corresponding
	# confidence scores
	(numRows, numCols) = scores.shape[2:4]
	rects = []
	confidences = []

	# loop over the number of rows
	for y in range(0, numRows):
		# extract the scores (probabilities), followed by the
		# geometrical data used to derive potential bounding box
		# coordinates that surround text
		scoresData = scores[0, 0, y]
		xData0 = geometry[0, 0, y]
		xData1 = geometry[0, 1, y]
		xData2 = geometry[0, 2, y]
		xData3 = geometry[0, 3, y]
		anglesData = geometry[0, 4, y]

		# loop over the number of columns
		for x in range(0, numCols):
			# if our score does not have sufficient probability,
			# ignore it
			if scoresData[x] < args["min_confidence"]:
				continue

			# compute the offset factor as our resulting feature
			# maps will be 4x smaller than the input image
			(offsetX, offsetY) = (x * 4.0, y * 4.0)

			# extract the rotation angle for the prediction and
			# then compute the sin and cosine
			angle = anglesData[x]
			cos = np.cos(angle)
			sin = np.sin(angle)

			# use the geometry volume to derive the width and height
			# of the bounding box
			h = xData0[x] + xData2[x]
			w = xData1[x] + xData3[x]

			# compute both the starting and ending (x, y)-coordinates
			# for the text prediction bounding box
			endX = int(offsetX + (cos * xData1[x]) + (sin * xData2[x]))
			endY = int(offsetY - (sin * xData1[x]) + (cos * xData2[x]))
			startX = int(endX - w)
			startY = int(endY - h)

			# add the bounding box coordinates and probability score
			# to our respective lists
			rects.append((startX, startY, endX, endY))
			confidences.append(scoresData[x])

	# return a tuple of the bounding boxes and associated confidences
	return (rects, confidences)
def GetDataForAnalyze():
	global lastAnalyze, googleSheetsService, spreadsheet,zonesCount,queue,responseStack,responseCount
	value_render_option = 'UNFORMATTED_VALUE'
	date_time_render_option = 'FORMATTED_STRING'
	majorDimension='COLUMNS'
	range_ = 'Tmp!A' + str(lastAnalyze + 1) + ':' + str(zonesCount+1) + str(currRow)
	response={}
	request = googleSheetsService.spreadsheets().values().get(spreadsheetId=spreadsheet['spreadsheetId'], range=range_,
															  majorDimension=majorDimension,
															  valueRenderOption=value_render_option,
															  dateTimeRenderOption=date_time_render_option)

	responseCount+=1
	requestNum=responseCount
	queue.append({'Number':responseCount,'Request':request,'Type':'get'})
	while True:
		f=False
		for responseFromStack in responseStack:
			if requestNum == responseFromStack['Number']:
				response = responseFromStack['Response']
				f=True
				break
		if f:
			break
	Timestamps=response['values'][0]
	Text=[]
	for i in range(1,len(response['values'])):
		Text.append((response['values'][i]))
	return Timestamps,Text

def Analyze(period):
	while listen:
		while run:
			global currRow,zonesCount
			time.sleep(round(period/2))
			global lastAnalyze,googleSheetsService,spreadsheet
			Result=[]
			Timestamps,Text = GetDataForAnalyze()
			for i in range(len(Text)):
				if Text[i]==[]:
					continue
				result = Text[i][0]
				pos = 0
				pos_r = 0
				column=[]
				startTimes = []
				endTimes = []
				ftime=str(Timestamps[0])
				for j in range(len(Text[i])-1):
					a = Text[i][j]
					b = Text[i][j+1]
					if (a==''):
						if (b!=''):
							result=b
						continue
					max_match = 0
					max_match_r = 0
					for match in difflib.SequenceMatcher(None, a, b).get_matching_blocks():
						if (match.size > max_match):
							max_match = match.size
							pos = match.b
					for match in difflib.SequenceMatcher(None, result, b).get_matching_blocks():
						if (match.size > max_match_r):
							max_match_r = match.size
							pos_r = match.a
					s = SequenceMatcher(lambda x: x == " ", a, b)
					if (s.ratio() < 0.3):  # NEW TEXT
						if(result!=''):
							if(ftime==Timestamps[j]):
								startTimes.append(ftime)
								column.append(result)
							else:
								startTimes.append(ftime)
								endTimes.append(str(Timestamps[j-1]))
								column.append(result)
						ftime=str(Timestamps[j])
						result = b
					else:
						result = result[:pos_r + max_match_r] + b[pos + max_match:]
				j=len(Text[i])-1
				if result!='':
					if (ftime == Timestamps[j]):
						startTimes.append(ftime)
						column.append(result)
					else:
						startTimes.append(ftime)
						endTimes.append(str(Timestamps[j - 1]))
						column.append(result)
				Result.append(startTimes)
				Result.append(endTimes)
				Result.append(column)
			lastAnalyze=lastAnalyze+len(Timestamps)-1
			WriteToResultSheet(Timestamps,Result)
			time.sleep((round(period/2)))
	return

def WriteToResultSheet(Timestamps,Result):
	global responseCount, responseStack
	startTime = str(Timestamps[0])
	endTime = str(Timestamps[len(Timestamps) - 1])
	lastRowResult = FindLastRow(googleSheetsService, spreadsheet['spreadsheetId'], 'Result')
	data = []
	tmp_dict = {}
	tmp_dict['range'] = 'Result!' + 'A' + str(lastRowResult + 1)
	tmp_dict['majorDimension'] = "ROWS"
	tmp_dict['values'] = [[startTime + '-' + endTime]]
	data.append(tmp_dict)
	for i in range(0, len(Result)):
		for j in range(0, len(Result[i])):
			tmp_dict = {}
			tmp_dict['range'] = 'Result!' + get_column_letter(i + 2) + str(lastRowResult + 2 + j)
			tmp_dict['majorDimension'] = "ROWS"
			tmp_dict['values'] = [[Result[i][j]]]
			data.append(tmp_dict)
	request = googleSheetsService.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet['spreadsheetId'],
															body={"valueInputOption": "USER_ENTERED", "data": data})

	responseCount += 1
	queue.append({'Number': responseCount, 'Request': request, 'Type': 'set'})
	return

def FindLastCol(service,spreadsheetId,shName):
	global responseCount,responseStack
	value_render_option = 'UNFORMATTED_VALUE'
	date_time_render_option = 'FORMATTED_STRING'
	range = shName+'!A1:CV10000'
	request = service.spreadsheets().values().get(spreadsheetId=spreadsheetId, range=range,
												  valueRenderOption=value_render_option,
												  dateTimeRenderOption=date_time_render_option)
	responseCount += 1
	requestNum = responseCount
	queue.append({'Number': responseCount, 'Request': request, 'Type': 'get'})
	while True:
		f = False
		for responseFromStack in responseStack:
			if requestNum == responseFromStack['Number']:
				response = responseFromStack['Response']
				responseStack.remove(responseFromStack)
				f = True
				break
		if f:
			break
	max_col = 1
	if 'values' in response:
		for row in response['values']:
			if len(row)>max_col:
				max_col = len(row)
	return get_column_letter(max_col)

def FindLastRow(service,spreadsheetId,shName):
	global responseCount,responseStack
	value_render_option = 'UNFORMATTED_VALUE'
	date_time_render_option = 'FORMATTED_STRING'
	range=shName+'!A1:CV10000'
	request = service.spreadsheets().values().get(spreadsheetId=spreadsheetId, range=range,
												  valueRenderOption=value_render_option,
												  dateTimeRenderOption=date_time_render_option)
	responseCount += 1
	requestNum = responseCount
	queue.append({'Number': responseCount, 'Request': request, 'Type': 'get'})
	while True:
		f = False
		for responseFromStack in responseStack:
			if requestNum == responseFromStack['Number']:
				response = responseFromStack['Response']
				responseStack.remove(responseFromStack)
				f = True
				break
		if f:
			break
	if 'values' in response:
		rowCount = len(response['values'])
	else:
		rowCount=1
	return rowCount



def video_to_text():
	# construct the argument parser and parse the arguments
	global run,listen,googleSheetsService, spreadsheet,args,resultStack,zonesCount,currRow,frameNum,frame,vs,capture
	while listen:
		# initialize the original frame dimensions, new frame dimensions,
		# and ratio between the dimensions
		(W, H) = (None, None)
		(newW, newH) = (args["width"], args["height"])
		(rW, rH) = (None, None)

		# define the two output layer names for the EAST detector model that
		# we are interested -- the first is the output probabilities and the
		# second can be used to derive the bounding box coordinates of text
		layerNames = [
			"feature_fusion/Conv_7/Sigmoid",
			"feature_fusion/concat_3"]

		# load the pre-trained EAST text detector
		print("[INFO] loading EAST text detector...")
		net = cv2.dnn.readNet(args["east"])

		# if a video path was not supplied, grab the reference to the web cam
		vs = cv2.VideoCapture(args["video"])
		print('vs is opened:',vs.isOpened())
		if not vs.isOpened():
			continue
		else:
			run=True
			capture=True
		fps = int(round(vs.get(cv2.CAP_PROP_FPS)))  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
		# loop over frames from the video stream
		zonesCount=0
		zones = []
		startTime=0
		endTime=1
		prevFrame=0
		dictToAnalyze={}
		dictToAnalyze['Timestamps']=[]
		dictToAnalyze['Text']=[]
		prevTime=0
		while (run):
			# grab the current frame, then handle if we are using a
			# VideoStream or VideoCapture object
			if frame is None:
				continue
			try:
				if not frame[0]:
					run=False
					break
			except Exception:
				continue
			#if (frameNum != prevFrame + round ((endTime-startTime)*fps))&(prevFrame!=0):
			#	continue
			startTime = time.time()
			frame = frame[1]
			# check to see if we have reached the end of the stream
			if frame is None:
				break
			capture=False
			try:
				frame = cv2.resize(frame, (newW, newH))
			except Exception:
				continue
			frame = np.array(Image.fromarray(frame).crop((0, round(2 / 3 * newH), newW, newH)))
			frame1=frame.copy()
			orig = frame1.copy()
			capture = True
			if W is None or H is None:
				(H, W) = frame1.shape[:2]
				rW = W / float(newW)
				rH = H / float(newH)
			# resize the frame, this time ignoring aspect ratio

			# construct a blob from the frame and then perform a forward pass
			# of the model to obtain the two output layer sets
			blob = cv2.dnn.blobFromImage(frame1, 1.0, (newW, newH),
				(123.68, 116.78, 103.94), swapRB=True, crop=False)
			net.setInput(blob)
			(scores, geometry) = net.forward(layerNames)

			# decode the predictions, then  apply non-maxima suppression to
			# suppress weak, overlapping bounding boxes
			(rects, confidences) = decode_predictions(scores, geometry,args)
			clusters=[]
			boxes = non_max_suppression(np.array(rects), probs=confidences)
			for (startX, startY, endX, endY) in boxes:
				startX = int(startX * rW)
				startY = int(startY * rH)
				endX = int(endX * rW)
				endY = int(endY * rH)
				clusterInfo = {}
				clusterInfo['middleY'] = (startY+endY)/2
				clusterInfo['middleX'] = (startX + endX) / 2
				clusterInfo['count'] = 1
				clusterInfo['startX'] = startX
				clusterInfo['startY'] = startY
				clusterInfo['endX'] = endX
				clusterInfo['endY'] = endY
				clusterInfo['avStrokeWidth'] = endY-startY
				clusterInfo['Zone']=''
				clusterInfo['Length']=clusterInfo['endX']-clusterInfo['startX']
				clusters.append(clusterInfo)
			flag = True
			for zone in zones:
				height=(zone['startY']+zone['endY'])/2
				zone['startY'] = zone['startY']+round(0.05*height)
				zone['endY'] = zone['endY'] - +round(0.05*height)
			while flag:
				flag = False
				for clust in clusters:
					for clustToPair in clusters:
						if (clust == clustToPair):
							continue
						smaller=0
						if ((clust['endX']-clust['startX'])<(clustToPair['endX']-clustToPair['startX'])):
							smaller = 1
						else:
							smaller = 2
						if smaller == 1:
							if (clust['startY']<clustToPair['middleY']<clust['endY'])&((abs(clust['startX']-clustToPair['endX'])<clust['endX']-clust['startX'])|(abs(clust['endX']-clustToPair['startX'])<clust['endX']-clust['startX'])|((clust['endX']<clustToPair['endX'])&(clust['startX']>clustToPair['startX'])))&(0.8*clust['avStrokeWidth']<clustToPair['avStrokeWidth']<1.2*clust['avStrokeWidth']):
								flag = True
								clust['startY'] = min(clust['startY'],clustToPair['startY'])
								clust['startX'] = min(clust['startX'], clustToPair['startX'])
								clust['endY'] = max(clust['endY'], clustToPair['endY'])
								clust['endX'] = max(clust['endX'], clustToPair['endX'])
								clust['middleY'] = (clust['startY'] + clust['endY']) / 2
								clust['middleX'] = (clust['startX'] + clust['endX']) / 2
								clust['avStrokeWidth'] = (clust['avStrokeWidth']*clust['count'] +clustToPair['avStrokeWidth'])/(clust['count']+1)
								clust['count'] += 1
								clust['Length'] = clust['endX']-clust['startX']
								clusters.remove(clustToPair)
						if smaller == 2:
							if (clust['startY']<clustToPair['middleY']<clust['endY'])&((abs(clustToPair['startX']-clust['endX'])<clustToPair['endX']-clustToPair['startX'])|(abs(clustToPair['endX']-clust['startX'])<clustToPair['endX']-clustToPair['startX'])|((clustToPair['endX']<clust['endX'])&(clustToPair['startX']>clust['startX'])))&(0.9*clustToPair['avStrokeWidth']<clust['avStrokeWidth']<1.2*clustToPair['avStrokeWidth']):
								flag = True
								clust['startY'] = min(clust['startY'],clustToPair['startY'])
								clust['startX'] = min(clust['startX'], clustToPair['startX'])
								clust['endY'] = max(clust['endY'], clustToPair['endY'])
								clust['endX'] = max(clust['endX'], clustToPair['endX'])
								clust['middleY'] = (clust['startY'] + clust['endY']) / 2
								clust['middleX'] = (clust['startX'] + clust['endX']) / 2
								clust['avStrokeWidth'] = (clustToPair['avStrokeWidth'] * clustToPair['count'] + clust['avStrokeWidth']) / (clustToPair['count'] + 1)
								clust['count'] += 1
								clust['Length'] = clust['endX'] - clust['startX']
								clusters.remove(clustToPair)
			clusters.sort(key=lambda x: x['Length'],reverse=True)
			for clust in clusters:
				newZone = True
				for zone in zones:
					if (zone['startY']<clust['middleY']<zone['endY']):
						if (zone['startX']<clust['middleX']<zone['endX'])|(SquareInteractPercent(zone['startX'],zone['startY'],zone['endX'],zone['endY'],clust['startX'],clust['startY'],clust['endX'],clust['endY'])>0.7):
							newZone = False
							clust['Zone']=zone['Zone']
							zone['startX'] = min(clust['startX'],zone['startX'])
							zone['startY'] = min(clust['startY'],zone['startY'])
							zone['endX'] =max(clust['endX'],zone['endX'])
							zone['endY'] = max(clust['endY'],zone['endY'])
							zone['avStrokeWidth'] = clust['avStrokeWidth']
							zone['Use'] = 'Yep'
				if newZone:
						zonesCount+=1
						zone = {}
						zone['startX']=clust['startX']
						zone['startY'] = clust['startY']
						zone['endX'] = clust['endX']
						zone['endY'] = clust['endY']
						zone['Zone'] = str(zonesCount)
						zone['avStrokeWidth'] = clust['avStrokeWidth']
						zone['Use'] = 'Yep'
						zones.append(zone)
						clust['Zone'] = zone['Zone']
			for zone in zones:
				for zoneToCompare in zones:
					if (zone != zoneToCompare):
						if (zone['startY']<(zoneToCompare['startY']+zoneToCompare['endY'])/2<zone['endY']):
							if (zone['startX'] < (zoneToCompare['startX']+zoneToCompare['endX'])/2 < zone['endX']) | (
									SquareInteractPercent(zone['startX'], zone['startY'], zone['endX'], zone['endY'],
														  zoneToCompare['startX'], zoneToCompare['startY'], zoneToCompare['endX'],
														  zoneToCompare['endY']) > 0.7):
								if ((zone['endX'] - zone['startX']) < (zoneToCompare['endX'] - zoneToCompare['startX'])):
									bigger = 2
								else:
									bigger = 1
								if bigger == 2:
									zone['startX'] = min(zoneToCompare['startX'], zone['startX'])
									zone['startY'] = min(zoneToCompare['startY'], zone['startY'])
									zone['endX'] = max(zoneToCompare['endX'], zone['endX'])
									zone['endY'] = max(zoneToCompare['endY'], zone['endY'])
									zone['Zone'] = min(zone['Zone'],zoneToCompare['Zone'])
									try:
										zones.remove(zoneToCompare)
									except Exception:
										continue
									zonesCount-=1
								else:
									zoneToCompare['startX'] = min(zoneToCompare['startX'], zone['startX'])
									zoneToCompare['startY'] = min(zoneToCompare['startY'], zone['startY'])
									zoneToCompare['endX'] = max(zoneToCompare['endX'], zone['endX'])
									zoneToCompare['endY'] = max(zoneToCompare['endY'], zone['endY'])
									zoneToCompare['Zone'] = min(zone['Zone'], zoneToCompare['Zone'])
									try:
										zones.remove(zone)
									except Exception:
										continue
									zonesCount -= 1
			textToSheet=[]
			timex=datetime.now().strftime('%H:%M:%S')
			for zone in zones:
				if (zone['Use']=='Yep'):
					imageCroped = Image.fromarray(orig).crop(
						(zone['startX']-5, zone['startY']-5, zone['endX']+5, zone['endY']+5))
					imageCroped.save('frame.jpg',dpi=(250,250))
					imageCroped.save('frames/'+str(timex)+'.'+str(zone['Zone'])+'.jpg')
					#os.system('tesseract frame.jpg output --oem 1 -l eng+rus')
					os.system('tesseract frame.jpg output --oem 1 -l eng tsv')
					text = ParseTessData()
					print(text)
					text = text.lstrip(" \n +-=)(|'")
					text = text.rstrip(" \n +-=)(|'")
					text = text.replace('\n', ' ').replace('\r', '')
					if text !='':
						tmp_dict = {}
						tmp_dict['Text'] = text
						tmp_dict['Zone'] = zone['Zone']
						textToSheet.append(tmp_dict)
			textToSheet.sort(key=lambda x: x['Zone'])
			writeToSheet(textToSheet, spreadsheet, googleSheetsService,currRow,timex)
			for zone in zones:
				zone['Use']=''
			currRow+=1
			endTime = time.time()

		# close all windows
	return

def VSCapture():
	while listen:
		global frame,capture,vs,frameNum,args
		while run:
			if (capture) & (vs is not None):
				if(vs.isOpened()):
					frame=vs.read()
					frameNum=vs.get(cv2.CAP_PROP_POS_FRAMES)
	return
def main():
	os.environ['OMP_THREAD_LIMIT'] = '4'
	global googleSheetsService, spreadsheet,run,listen,currRow,zonesCount,queue,responseCount,responseStack,args,lastAnalyze,frame,capture,vs,frameNum
	run = False
	frame=None
	listen = True
	capture=False
	vs=None
	zonesCount = 0
	responseCount = 0
	queue = deque()
	responseStack = []
	googleSheetsService, spreadsheet = initOutputSheet()
	spreadsheet['spreadsheetId'] = '1F749DStmew6FbBm9XKPSeureHelXr1vnJmXdcZtX3xc'
	googleRequestQueueThread = Thread(target=GoogleRequestQueue)
	googleRequestQueueThread.start()
	lastAnalyze = FindLastRow(googleSheetsService, spreadsheet['spreadsheetId'], 'Tmp')
	currRow = lastAnalyze
	args = {}
	args['east'] = 'frozen_east_text_detection.pb'
	args["video"] = 'rtmp://hse.auditory.ru/live'
	args['type'] = ''
	args['min_confidence'] = 0.5
	args['width'] = 736
	args['height'] = 576
	videoToTextThread = Thread(target=video_to_text)
	analyzerThread = Thread(target=Analyze,args=[60])
	videoCaptureThread = Thread(target=VSCapture)
	analyzerThread.start()
	videoCaptureThread.start()
	video_to_text().start()
	videoToTextThread.join()
	analyzerThread.join()
	videoCaptureThread.join()
	googleRequestQueueThread.join()

main()
